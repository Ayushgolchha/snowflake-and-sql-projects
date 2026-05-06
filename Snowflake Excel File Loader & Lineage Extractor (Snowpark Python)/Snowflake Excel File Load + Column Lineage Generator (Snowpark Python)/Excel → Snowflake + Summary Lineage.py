#########LOAD EXCEL FILE & CREATE A DATA LINEAGE INTO SNOWFLAKE


import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from snowflake.snowpark.files import SnowflakeFile


def sanitize_name(name: str) -> str:
    name = str(name).strip().upper()
    name = re.sub(r"[^A-Z0-9]+", "_", name)
    name = re.sub(r"_+", "_", name)
    return name.strip("_")


def is_summary_sheet(sheet_name: str) -> bool:
    return "SUMMARY" in sanitize_name(sheet_name)


def extract_sheet_references(formula: str):
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return []

    pattern = re.compile(
        r"(?:'([^']+)'|([A-Za-z0-9_]+))!"
        r"(\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)"
    )

    refs = []
    for match in pattern.finditer(formula):
        sheet_name = match.group(1) or match.group(2)
        cell_range = match.group(3)
        refs.append({
            "referenced_sheet_original": str(sheet_name).strip(),
            "referenced_sheet_table": sanitize_name(sheet_name),
            "referenced_range": cell_range
        })

    return refs


def get_headers_map(ws):
    headers_map = {}
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        header_value = ws.cell(row=1, column=col_idx).value
        headers_map[col_letter] = sanitize_name(header_value) if header_value is not None else f"COL_{col_letter}"
    return headers_map


def extract_columns_from_range(cell_range: str):
    parts = cell_range.replace("$", "").split(":")
    start_ref = parts[0]
    end_ref = parts[-1]

    start_col = re.match(r"([A-Z]+)", start_ref).group(1)
    end_col = re.match(r"([A-Z]+)", end_ref).group(1)

    start_idx = column_index_from_string(start_col)
    end_idx = column_index_from_string(end_col)

    return [get_column_letter(i) for i in range(start_idx, end_idx + 1)]


def worksheet_to_records(ws):
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []

    raw_headers = all_rows[0]
    headers = []
    for idx, h in enumerate(raw_headers, start=1):
        headers.append(sanitize_name(h) if h is not None else f"COL_{get_column_letter(idx)}")

    records = []
    for row in all_rows[1:]:
        if row is None or all(v is None for v in row):
            continue
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = value
        records.append(row_dict)

    return records


def main(session):
    file_path = "@EXCEL_FILES_LOAD.EXCEL_FILES.EXCEL_FILES_STAGE/Test/Test_Excel_Load_SF.xlsx"
    target_db = "EXCEL_FILES_LOAD"
    target_schema = "EXCEL_FILES"
    lineage_table_name = [target_db, target_schema, "SUMMARY_LINEAGE_COLUMNS"]

    # FIX: allow owner-controlled stage path
    with SnowflakeFile.open(file_path, "rb", require_scoped_url=False) as f:
        file_bytes = f.read()

    wb_formula = load_workbook(BytesIO(file_bytes), data_only=False)

    sheet_headers = {}
    load_summary = []

    for sheet_name in wb_formula.sheetnames:
        if is_summary_sheet(sheet_name):
            continue

        ws = wb_formula[sheet_name]
        table_name = sanitize_name(sheet_name)

        records = worksheet_to_records(ws)
        sheet_headers[table_name] = get_headers_map(ws)

        if not records:
            load_summary.append({
                "sheet_name": sheet_name,
                "table_name": table_name,
                "row_count": 0,
                "status": "SKIPPED_EMPTY"
            })
            continue

        df = session.create_dataframe(records)
        df.write.save_as_table([target_db, target_schema, table_name], mode="overwrite")

        load_summary.append({
            "sheet_name": sheet_name,
            "table_name": table_name,
            "row_count": len(records),
            "status": "LOADED"
        })

    lineage_rows = []

    for summary_sheet_name in wb_formula.sheetnames:
        if not is_summary_sheet(summary_sheet_name):
            continue

        summary_ws = wb_formula[summary_sheet_name]

        for row_idx in range(1, summary_ws.max_row + 1):
            formula_value = summary_ws[f"B{row_idx}"].value

            if not isinstance(formula_value, str) or not formula_value.startswith("="):
                continue

            refs = extract_sheet_references(formula_value)

            for ref in refs:
                referenced_sheet_table = ref["referenced_sheet_table"]
                referenced_range = ref["referenced_range"]
                header_map = sheet_headers.get(referenced_sheet_table, {})
                excel_cols = extract_columns_from_range(referenced_range)

                for excel_col in excel_cols:
                    lineage_rows.append({
                        "PARENT_TABLE": "SUMMARY",
                        "REFERENCE_TABLE": referenced_sheet_table,
                        "COLUMNS_USE": header_map.get(excel_col, f"COL_{excel_col}")
                    })

    deduped = []
    seen = set()
    for row in lineage_rows:
        key = (row["PARENT_TABLE"], row["REFERENCE_TABLE"], row["COLUMNS_USE"])
        if key not in seen:
            seen.add(key)
            deduped.append(row)

    if deduped:
        lineage_df = session.create_dataframe(deduped)
        lineage_df.write.save_as_table(lineage_table_name, mode="overwrite")

    return {
        "sheet_load_summary": load_summary,
        "lineage_count": len(deduped),
        "lineage_table": ".".join(lineage_table_name)
    }
